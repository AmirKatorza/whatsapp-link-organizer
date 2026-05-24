import csv
import os
from collections import defaultdict
from html import escape
from typing import Iterable

from src.models import ExtractedLink

_HTML_HEADER = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>WhatsApp Link Organizer</title>
<style>
  body { font-family: system-ui, sans-serif; margin: 0; padding: 1rem 2rem; background: #f8f9fa; color: #212529; }
  h1 { font-size: 1.6rem; margin-bottom: 0.25rem; }
  .summary { color: #6c757d; margin-top: 0; margin-bottom: 1.5rem; }
  nav ul { list-style: none; padding: 0; display: flex; flex-wrap: wrap; gap: 0.5rem; margin-bottom: 2rem; }
  nav li a { background: #e9ecef; border-radius: 4px; padding: 0.3rem 0.7rem; text-decoration: none; color: #495057; font-size: 0.9rem; }
  nav li a:hover { background: #dee2e6; }
  section { margin-bottom: 3rem; }
  h2 { font-size: 1.2rem; border-bottom: 2px solid #dee2e6; padding-bottom: 0.4rem; }
  .count { font-weight: normal; color: #6c757d; font-size: 1rem; }
  table { width: 100%; border-collapse: collapse; background: #fff; border-radius: 6px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,.08); }
  thead { background: #343a40; color: #fff; }
  th { padding: 0.6rem 0.8rem; text-align: left; font-size: 0.85rem; font-weight: 600; }
  td { padding: 0.5rem 0.8rem; font-size: 0.85rem; border-bottom: 1px solid #f1f3f5; vertical-align: top; }
  tr:last-child td { border-bottom: none; }
  tr:hover td { background: #f8f9fa; }
  .ts { white-space: nowrap; color: #6c757d; width: 11rem; }
  .url a { color: #0d6efd; word-break: break-all; }
  .conf { text-align: right; width: 4rem; color: #6c757d; }
</style>
</head>
<body>"""

_HTML_FOOTER = "</body></html>"


def write_html_report(csv_path: str, output_path: str) -> None:
    groups: dict[str, list[dict]] = defaultdict(list)

    with open(csv_path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            category = row.get("category") or "Other / Review"
            groups[category].append(row)

    def _sort_key(item: tuple[str, list]) -> tuple[int, int]:
        cat, rows = item
        return (1, -len(rows)) if cat == "Other / Review" else (0, -len(rows))

    sorted_groups = sorted(groups.items(), key=_sort_key)
    total = sum(len(rows) for rows in groups.values())

    parts = [_HTML_HEADER]
    parts.append(f'<h1>WhatsApp Link Organizer</h1>')
    parts.append(f'<p class="summary">{total} links across {len(groups)} categories</p>')

    parts.append('<nav><ul>')
    for cat, rows in sorted_groups:
        anchor = _anchor(cat)
        parts.append(f'<li><a href="#{anchor}">{escape(cat)} ({len(rows)})</a></li>')
    parts.append('</ul></nav>')

    for cat, rows in sorted_groups:
        anchor = _anchor(cat)
        parts.append(f'<section id="{anchor}">')
        parts.append(f'<h2>{escape(cat)} <span class="count">({len(rows)})</span></h2>')
        parts.append('<table><thead><tr><th>Date</th><th>Link</th><th>Title / Context</th><th>Conf.</th></tr></thead><tbody>')
        for row in rows:
            timestamp = escape(row.get("timestamp") or "")
            url = row.get("url") or ""
            title = row.get("title") or ""
            context = row.get("context") or row.get("message_text") or ""
            confidence = row.get("category_confidence") or ""
            display = title if title else (context[:80] + "…" if len(context) > 80 else context)
            short_url = url[:60] + ("…" if len(url) > 60 else "")
            parts.append(
                f'<tr>'
                f'<td class="ts">{timestamp}</td>'
                f'<td class="url"><a href="{escape(url)}" target="_blank" rel="noopener">{escape(short_url)}</a></td>'
                f'<td>{escape(display)}</td>'
                f'<td class="conf">{escape(confidence)}</td>'
                f'</tr>'
            )
        parts.append('</tbody></table></section>')

    parts.append(_HTML_FOOTER)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(parts))


def _anchor(text: str) -> str:
    import re
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def write_review_spreadsheet(csv_path: str, output_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    CATEGORIES = [
        "SMC / Trading Concepts",
        "General Trading / Investing",
        "Daily Affirmations / Mindset",
        "Psychology / Self Development",
        "Healthy Cooking / Recipes",
        "Nutrition / Health",
        "Fitness / Training",
        "Business / Entrepreneurship",
        "Other / Review",
    ]

    COLUMNS = [
        ("Date",            "timestamp",              18),
        ("Platform",        "platform",               12),
        ("URL",             "url",                    10),
        ("Title",           "title",                  30),
        ("Context",         "context",                45),
        ("AI Category",     "category",               22),
        ("Confidence",      "category_confidence",    11),
        ("Review Category", None,                     22),
        ("Notes",           None,                     40),
        ("Reason",          "categorization_reason",  45),
    ]

    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = sorted(csv.DictReader(f), key=lambda r: r.get("timestamp") or "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Review"

    header_fill = PatternFill("solid", fgColor="1F3864")
    other_fill  = PatternFill("solid", fgColor="FFF9C4")
    header_font = Font(bold=True, color="FFFFFF")
    wrap        = Alignment(wrap_text=True, vertical="top")

    # Header row
    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "B2"  # freeze row 1 + column A

    # Data validation dropdown for "Review Category" (column H = index 8)
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(CATEGORIES) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)

    for row_idx, row in enumerate(rows, start=2):
        is_other = (row.get("category") or "") == "Other / Review"
        row_fill = other_fill if is_other else None

        for col_idx, (_, field, _) in enumerate(COLUMNS, start=1):
            if field == "url":
                url_val = row.get("url") or ""
                cell = ws.cell(row=row_idx, column=col_idx,
                               value=f'=HYPERLINK("{url_val}","open")')
            elif field is None:
                cell = ws.cell(row=row_idx, column=col_idx, value="")
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(field) or "")

            cell.alignment = wrap
            if row_fill:
                cell.fill = row_fill

        # Attach dropdown to Review Category cell (column H)
        dv.add(ws.cell(row=row_idx, column=8))

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)


def write_links_to_csv(records: Iterable[ExtractedLink], output_path: str) -> None:
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "timestamp",
                "sender",
                "message_text",
                "url",
                "platform",
                "context",
            ]
        )

        for r in records:
            writer.writerow(
                [
                    r.timestamp,
                    r.sender,
                    r.message_text,
                    r.url,
                    r.platform,
                    r.context or "",
                ]
            )